from datetime import datetime, timedelta
from typing import List, Set, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import random
from app.models import Employee, Skill, ShiftType, Task, DayOffRequest, ScheduleSettings

class Shift:
    def __init__(self, date: datetime, is_morning: bool):
        self.date = date
        self.is_morning = is_morning
        self.start_time = datetime.combine(date, datetime.min.time().replace(hour=9 if is_morning else 16))
        self.end_time = datetime.combine(date, datetime.min.time().replace(hour=16 if is_morning else 22))

class Task:
    def __init__(self, shift: Shift, task_type: str):
        self.shift = shift
        self.task_type = task_type

class Assignment:
    def __init__(self, task: Task, employee: Employee = None):
        self.task = task
        self.employee = employee

class ScheduleSolution:
    def __init__(self, assignments: List[Assignment], employees: List[Employee]):
        self.assignments = assignments
        self.employees = employees

def create_problem(start_date: datetime, employees: List[Employee], day_off_requests: Dict[str, List[datetime]]):
    assignments = []
    for day in range(21):
        current_date = start_date + timedelta(days=day)
        for is_morning in [True, False]:
            shift = Shift(current_date, is_morning)
            for task_type in ['cold', 'expo', 'grill']:
                task = Task(shift, task_type)
                assignments.append(Assignment(task))
    return ScheduleSolution(assignments, employees)

def solve_schedule(solution: ScheduleSolution, day_off_requests: Dict[str, List[datetime]]):
    employee_schedules = {e.name: [] for e in solution.employees}
    employee_hours = {e.name: 0 for e in solution.employees}
    employee_weekly_workdays = {e.name: {0: 0, 1: 0, 2: 0} for e in solution.employees}  # Track workdays for 3 weeks

    for assignment in solution.assignments:
        date = assignment.task.shift.date
        shift_duration = (assignment.task.shift.end_time - assignment.task.shift.start_time).total_seconds() / 3600
        week_number = (date - solution.assignments[0].task.shift.date).days // 7

        qualified_employees = [
            e for e in solution.employees 
            if assignment.task.task_type in e.skills 
            and date not in day_off_requests.get(e.name, [])
            and len([s for s in employee_schedules[e.name] if s.date == date]) == 0
            and len([s for s in employee_schedules[e.name] if (date - s.date).days < 5]) < 5
            and employee_hours[e.name] + shift_duration <= 120
            and employee_weekly_workdays[e.name][week_number] < 5  # Ensure max 5 workdays per week
        ]

        if not qualified_employees:
            continue

        # Prioritize employees with fewer hours and fewer workdays this week
        qualified_employees.sort(key=lambda e: (employee_hours[e.name], employee_weekly_workdays[e.name][week_number]))

        # Ensure at least one supervisor per shift
        supervisors = [e for e in qualified_employees if e.is_supervisor]
        if not any(e.is_supervisor for e in [a.employee for a in solution.assignments if a.task.shift == assignment.task.shift and a.employee]):
            qualified_employees = supervisors if supervisors else qualified_employees

        if qualified_employees:
            selected_employee = qualified_employees[0]
            assignment.employee = selected_employee
            employee_schedules[selected_employee.name].append(assignment.task.shift)
            employee_hours[selected_employee.name] += shift_duration
            employee_weekly_workdays[selected_employee.name][week_number] += 1

    return solution

def create_excel_output(solution: ScheduleSolution, start_date: datetime, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Write headers
    ws.cell(row=1, column=1, value="Employee")
    for i in range(21):
        date = start_date + timedelta(days=i)
        ws.cell(row=1, column=i+2, value=f"{date.strftime('%Y-%m-%d')} ({date.strftime('%a')})")
    ws.cell(row=1, column=23, value="Total Hours")

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Write employee data
    employee_hours = {employee.name: 0 for employee in solution.employees}
    for i, employee in enumerate(solution.employees):
        ws.cell(row=i+2, column=1, value=employee.name)
        for j in range(21):
            date = start_date + timedelta(days=j)
            cell = ws.cell(row=i+2, column=j+2)
            assignments = [a for a in solution.assignments if a.employee == employee and a.task.shift.date == date]
            if assignments:
                task_info = []
                for assignment in assignments:
                    task = assignment.task
                    task_info.append(f"{task.task_type}: {task.shift.start_time.strftime('%H:%M')}-{task.shift.end_time.strftime('%H:%M')}")
                    employee_hours[employee.name] += (task.shift.end_time - task.shift.start_time).total_seconds() / 3600
                cell.value = "\n".join(task_info)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        ws.cell(row=i+2, column=23, value=f"{employee_hours[employee.name]:.2f}")

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)

def generate_schedule():
    employees = Employee.query.all()
    settings = ScheduleSettings.query.order_by(ScheduleSettings.id.desc()).first()
    
    if not settings:
        raise ValueError("Schedule settings not found. Please set up schedule settings first.")

    start_date = settings.roster_start
    employee_list = [Employee(e.name, {s.name for s in e.skills}, e.is_supervisor) for e in employees]
    
    day_off_requests = {}
    for request in DayOffRequest.query.all():
        if request.employee.name not in day_off_requests:
            day_off_requests[request.employee.name] = []
        day_off_requests[request.employee.name].append(request.date)

    problem = create_problem(start_date, employee_list, day_off_requests)
    solution = solve_schedule(problem, day_off_requests)

    # Create Excel output
    output_dir = os.path.join(os.getcwd(), 'schedules')
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"schedule_{start_date.strftime('%Y-%m-%d')}.xlsx")
    create_excel_output(solution, start_date, output_file)

    return output_file